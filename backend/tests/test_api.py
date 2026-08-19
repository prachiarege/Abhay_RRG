"""End-to-end API tests: CSV files -> ingestion -> engine -> HTTP response.

Runs against an isolated temporary SQLite database and the CSV provider, so the suite is
deterministic and needs no network. Environment variables are set BEFORE importing the
app because `app.db` builds its engine at import time from the resolved settings.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

_TMP = Path(tempfile.mkdtemp(prefix="rrg_api_test_"))
_CSV_DIR = _TMP / "csv"
_CSV_DIR.mkdir(parents=True, exist_ok=True)

os.environ["RRG_DATABASE_URL"] = f"sqlite:///{(_TMP / 'test.db').as_posix()}"
os.environ["RRG_DATA_PROVIDER"] = "csv"
os.environ["RRG_CSV_DATA_DIR"] = str(_CSV_DIR)
os.environ["RRG_AUTO_REFRESH_ENABLED"] = "false"
os.environ["RRG_RATE_LIMIT_PER_MINUTE"] = "0"  # disabled; the limiter has its own test

from fastapi.testclient import TestClient  # noqa: E402

from app.db import init_db, session_scope  # noqa: E402
from app.main import app  # noqa: E402
from app.seed import seed_universe  # noqa: E402
from app.services.ingestion import refresh_prices  # noqa: E402

# The ten sectors seeded as the default on-screen universe, plus the default benchmark.
FIXTURE_SYMBOLS = {
    "CRSLDX": 0.00035,      # NIFTY 500 benchmark
    "CNXIT": 0.00055,
    "CNXAUTO": 0.00040,
    "NSEBANK": 0.00045,
    "CNXFMCG": 0.00030,
    "CNXPHARMA": 0.00025,
    "CNXMETAL": 0.00015,
    "CNXREALTY": 0.00060,
    "CNXMEDIA": -0.00020,
    "CNXENERGY": 0.00033,
    "CNXINFRA": 0.00038,
}

BARS = 1500  # ~6 years of business days: comfortably past any warm-up requirement


def _write_fixture_csvs() -> None:
    """Deterministic synthetic daily series, one CSV per provider symbol."""
    index = pd.bdate_range(end="2026-08-14", periods=BARS, name="date")
    for position, (symbol, drift) in enumerate(FIXTURE_SYMBOLS.items()):
        generator = np.random.default_rng(1000 + position)
        shocks = generator.normal(loc=drift, scale=0.009, size=BARS)
        # A slow cycle per sector, phase-shifted, so sectors genuinely rotate between
        # quadrants instead of all drifting in step.
        cycle = 0.12 * np.sin(np.linspace(0, 4 * np.pi, BARS) + position)
        close = 10_000.0 * np.exp(np.cumsum(shocks) + cycle)
        frame = pd.DataFrame(
            {
                "Date": index.strftime("%Y-%m-%d"),
                "Open": close,
                "High": close * 1.004,
                "Low": close * 0.996,
                "Close": close,
                "Volume": 1_000_000,
            }
        )
        frame.to_csv(_CSV_DIR / f"{symbol}.csv", index=False)


@pytest.fixture(scope="module", autouse=True)
def _prepared_database():
    _write_fixture_csvs()
    init_db()
    with session_scope() as session:
        seed_universe(session)
        session.commit()
        # Sectors without a fixture CSV fail here on purpose: that exercises the
        # failure-isolation path (SRS 46) rather than hiding it.
        refresh_prices(session, trigger="test")
    yield


@pytest.fixture(scope="module")
def client(_prepared_database) -> TestClient:
    with TestClient(app) as test_client:
        yield test_client


# ------------------------------------------------------------------------------ metadata


def test_health(client: TestClient):
    response = client.get("/api/health")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "ok"
    assert body["provider"] == "csv"
    assert body["data"]["latest_date"] is not None


def test_sectors_listed_from_database(client: TestClient):
    response = client.get("/api/sectors")
    assert response.status_code == 200
    sectors = response.json()
    assert len(sectors) == 17, "all 17 SRS 2.1 sectors should be present as rows"
    assert sum(1 for s in sectors if s["is_default"]) == 10
    unavailable = [s["symbol"] for s in sectors if not s["available"]]
    assert set(unavailable) == {
        "NIFTY_OIL_GAS",
        "NIFTY_CONSUMER_DUR",
        "NIFTY_HEALTHCARE",
    }


def test_benchmarks_listed(client: TestClient):
    body = client.get("/api/benchmarks").json()
    symbols = {b["symbol"] for b in body}
    assert "NIFTY500" in symbols
    assert next(b for b in body if b["symbol"] == "NIFTY500")["is_default"] is True


def test_config_exposes_defaults(client: TestClient):
    body = client.get("/api/config").json()
    assert body["defaults"]["benchmark"] == "NIFTY500"
    assert body["defaults"]["center"] == 100.0
    assert body["quadrants"] == ["Leading", "Weakening", "Lagging", "Improving"]


# ----------------------------------------------------------------------------------- rrg


def test_rrg_default_request(client: TestClient):
    response = client.get("/api/rrg")
    assert response.status_code == 200, response.text
    body = response.json()

    assert body["benchmark"] == "NIFTY500"
    assert body["frequency"] == "weekly"
    assert body["tail_length"] == 10
    assert body["center"] == 100.0
    assert body["engine_version"] == "1.0.0"
    assert len(body["params_fingerprint"]) == 16
    assert len(body["sectors"]) == 10

    for sector in body["sectors"]:
        assert sector["quadrant"] in ("Leading", "Weakening", "Lagging", "Improving")
        assert len(sector["tail"]) == 10
        assert sector["tail"][-1]["date"] == sector["date"]
        assert sector["rs_ratio"] is not None
        assert sector["rs_momentum"] is not None
        assert 0 <= sector["rotation_score"] <= 100


def test_quadrant_matches_coordinates(client: TestClient):
    """The API's quadrant label must agree with its own numbers (SRS 12 orientation)."""
    body = client.get("/api/rrg").json()
    centre = body["center"]
    for sector in body["sectors"]:
        x, y = sector["rs_ratio"], sector["rs_momentum"]
        expected = {
            (True, True): "Leading",
            (True, False): "Weakening",
            (False, False): "Lagging",
            (False, True): "Improving",
        }[(x >= centre, y >= centre)]
        assert sector["quadrant"] == expected, f"{sector['symbol']} at ({x}, {y})"


def test_unavailable_sectors_do_not_break_the_response(client: TestClient):
    """Requesting a sector with no data returns the others, not an error (SRS 46)."""
    response = client.get("/api/rrg", params={"sectors": "NIFTY_IT,NIFTY_OIL_GAS"})
    assert response.status_code == 200
    body = response.json()
    assert [s["symbol"] for s in body["sectors"]] == ["NIFTY_IT"]
    assert [u["symbol"] for u in body["unavailable"]] == ["NIFTY_OIL_GAS"]


def test_tail_length_respected(client: TestClient):
    for tail in (5, 15, 30):
        body = client.get("/api/rrg", params={"tail": tail}).json()
        assert body["tail_length"] == tail
        assert all(len(s["tail"]) == tail for s in body["sectors"])


def test_daily_and_weekly_differ(client: TestClient):
    weekly = client.get("/api/rrg", params={"frequency": "weekly"}).json()
    daily = client.get("/api/rrg", params={"frequency": "daily"}).json()
    assert weekly["bars_available"] < daily["bars_available"]
    weekly_it = next(s for s in weekly["sectors"] if s["symbol"] == "NIFTY_IT")
    daily_it = next(s for s in daily["sectors"] if s["symbol"] == "NIFTY_IT")
    assert weekly_it["rs_ratio"] != daily_it["rs_ratio"]


def test_historical_date_reproduces_the_past(client: TestClient):
    """AC-14: a historical request must return that date, computed without later data."""
    dates = client.get("/api/rrg/dates").json()["dates"]
    target = dates[-20]

    first = client.get("/api/rrg", params={"as_of": target}).json()
    assert first["date"] == target
    assert first["requested_as_of"] == target

    # Same request again must be identical (SRS 50.2), and must not equal "latest".
    second = client.get("/api/rrg", params={"as_of": target}).json()
    assert first["sectors"] == second["sectors"]

    latest = client.get("/api/rrg").json()
    assert latest["date"] != target


def test_playback_dates_exclude_warmup(client: TestClient):
    body = client.get("/api/rrg/dates").json()
    assert body["count"] > 50
    assert body["warmup_bars"] == 41
    rrg = client.get("/api/rrg").json()
    assert body["count"] == rrg["bars_available"] - body["warmup_bars"] + 1


def test_insufficient_history_is_a_clear_error(client: TestClient):
    """Asking for more tail than the data supports must explain itself, not guess."""
    response = client.get(
        "/api/rrg", params={"frequency": "weekly", "tail": 60, "rs_period": 250}
    )
    assert response.status_code == 409
    detail = response.json()["detail"]
    assert "warm-up" in detail and "needs" in detail


def test_invalid_parameters_rejected(client: TestClient):
    assert client.get("/api/rrg", params={"frequency": "monthly"}).status_code == 422
    assert client.get("/api/rrg", params={"tail": 0}).status_code == 422
    assert client.get("/api/rrg", params={"rs_period": 1}).status_code == 422
    assert client.get("/api/rrg", params={"smoothing_method": "kalman"}).status_code == 422


def test_unknown_benchmark_is_400(client: TestClient):
    response = client.get("/api/rrg", params={"benchmark": "NOT_A_BENCHMARK"})
    assert response.status_code == 400


def test_sector_detail(client: TestClient):
    body = client.get("/api/sectors/NIFTY_IT/detail").json()
    assert body["symbol"] == "NIFTY_IT"
    assert body["quadrant"] in ("Leading", "Weakening", "Lagging", "Improving")
    assert len(body["history"]) > 100
    assert set(body["relative_returns"]) == {"1d", "1w", "1m", "3m", "6m", "1y"}
    assert body["direction_label"]


def test_sector_detail_unknown_symbol(client: TestClient):
    assert client.get("/api/sectors/NOPE/detail").status_code == 404


# -------------------------------------------------------------------------------- export


def test_csv_export_matches_screen_values(client: TestClient):
    """SRS 52.8: exported values must equal what the chart shows."""
    payload = client.get("/api/rrg").json()
    response = client.get("/api/export/rrg.csv")
    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]

    import csv as csv_module
    import io

    rows = list(csv_module.DictReader(io.StringIO(response.text)))
    assert rows

    for sector in payload["sectors"]:
        head = next(
            r
            for r in rows
            if r["symbol"] == sector["symbol"] and r["is_latest"] == "True"
        )
        assert float(head["rs_ratio"]) == sector["rs_ratio"]
        assert float(head["rs_momentum"]) == sector["rs_momentum"]
        assert head["quadrant"] == sector["quadrant"]
        assert float(head["rotation_score"]) == sector["rotation_score"]


def test_xlsx_export(client: TestClient):
    response = client.get("/api/export/rrg.xlsx")
    assert response.status_code == 200
    assert response.content[:2] == b"PK"  # a real xlsx is a zip archive

    import io

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["RRG Values", "Parameters"]
    assert workbook["RRG Values"].max_row > 10
    # The parameter sheet must record enough to reproduce the export later.
    recorded = {row[0] for row in workbook["Parameters"].iter_rows(values_only=True)}
    assert "Engine version" in recorded
    assert "Parameter fingerprint" in recorded


# --------------------------------------------------------------------------------- admin


def test_manual_refresh(client: TestClient):
    response = client.post("/api/refresh")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] in ("success", "partial")
    assert body["succeeded"] >= len(FIXTURE_SYMBOLS)
    assert "cache_entries_cleared" in body


def test_refresh_rejects_unknown_symbols(client: TestClient):
    response = client.post("/api/refresh", json={"symbols": ["NOT_REAL"]})
    assert response.status_code == 400


def test_cache_hit_is_recorded(client: TestClient):
    client.post("/api/admin/cache/clear")
    params = {"benchmark": "NIFTY500", "frequency": "weekly", "tail": 12}
    client.get("/api/rrg", params=params)
    before = client.get("/api/health").json()["cache"]["hits"]
    client.get("/api/rrg", params=params)
    after = client.get("/api/health").json()["cache"]["hits"]
    assert after > before
